from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def format(file_name):    
    # Load the existing Excel workbook
    wb = load_workbook(file_name)

    # Define styles once
    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    currency_fmt = "$#,##0.00"

    # Loop through every sheet in the workbook
    for ws in wb.worksheets:
        # Make headers bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Adjust column widths based on header
        for col in ws.iter_cols(1, ws.max_column):
            header_value = col[0].value
            if header_value == "Date":
                ws.column_dimensions[col[0].column_letter].width = 15
            elif header_value == "Transaction Description":
                ws.column_dimensions[col[0].column_letter].width = 60
            elif header_value in ["Debit/Cheque", "Credit/Deposit"]:
                ws.column_dimensions[col[0].column_letter].width = 15
            elif header_value == "Balance":
                ws.column_dimensions[col[0].column_letter].width = 15

        # Currency formatting for debit, credit, and balance
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx, cell in enumerate(row, start=1):
                header = ws.cell(1, col_idx).value
                if header in ["Debit/Cheque", "Credit/Deposit", "Balance"] and cell.value not in ("", None):
                    cell.number_format = currency_fmt

        # Alternating row color & black borders
        for row in range(2, ws.max_row + 1):
            for cell in ws[row]:
                if row % 2 == 0:
                    cell.fill = fill_gray
                cell.border = thin_border

    # Save changes after processing all sheets
    wb.save(file_name)
    wb.close()